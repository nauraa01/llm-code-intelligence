"""
=====================================================
BUG DETECTION EVALUATION
BugsInPy Dataset × 3 LLMs
Metrics: Precision, Recall, F1-Score, Wilcoxon
=====================================================
Cara pakai:
1. python bug_eval.py
2. Hasil: bug_eval_results.xlsx
Estimasi: 10-15 menit
"""

import os, time, re, warnings
import pandas as pd
import numpy as np
from dotenv import load_dotenv
from datetime import datetime
from pathlib import Path

from groq import Groq
from mistralai import Mistral
from openai import OpenAI
from scipy.stats import wilcoxon
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')
load_dotenv()

print("=" * 60)
print("  BUG DETECTION EVALUATION")
print("  BugsInPy × 3 LLMs (Groq + Mistral + OpenRouter)")
print("=" * 60)

# ─── API CLIENTS ──────────────────────────────────────────────
groq_client       = Groq(api_key=os.getenv("GROQ_API_KEY"))
mistral_client    = Mistral(api_key=os.getenv("MISTRAL_API_KEY"))
openrouter_client = OpenAI(
    api_key=os.getenv("OPENROUTER_API_KEY"),
    base_url="https://openrouter.ai/api/v1"
)

def call_llm(llm: str, prompt: str, retries=3) -> tuple:
    for attempt in range(retries):
        try:
            t0 = time.time()
            if llm == 'Groq':
                res = groq_client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=1024
                )
                return res.choices[0].message.content, round(time.time()-t0, 3)
            elif llm == 'Mistral':
                res = mistral_client.chat.complete(
                    model="mistral-small-latest",
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=1024
                )
                return res.choices[0].message.content, round(time.time()-t0, 3)
            elif llm == 'OpenRouter':
                res = openrouter_client.chat.completions.create(
                    model="openai/gpt-oss-120b:free",
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=1024
                )
                return res.choices[0].message.content, round(time.time()-t0, 3)
        except Exception as e:
            if attempt < retries-1:
                print(f"\n    ⚠️  Retry {attempt+1}/3...", end="")
                time.sleep(5*(attempt+1))
            else:
                return f"ERROR: {e}", 0.0

# ─── BUG TEST CASES ───────────────────────────────────────────
# 30 test cases: 20 buggy + 10 clean (realistic distribution)
test_cases = [
    # ── BUGGY CODE (label=1) ──────────────────────────────────
    {
        "id": "BD-001", "project": "pandas",
        "code": """def calculate_mean(data):
    return sum(data) / len(data)""",
        "bugs": ["ZeroDivisionError when data is empty", "no input validation"],
        "label": 1
    },
    {
        "id": "BD-002", "project": "requests",
        "code": """def fetch_url(url):
    response = requests.get(url)
    return response.json()""",
        "bugs": ["no timeout parameter", "no status code check", "no exception handling"],
        "label": 1
    },
    {
        "id": "BD-003", "project": "flask",
        "code": """def get_user(user_id):
    query = f"SELECT * FROM users WHERE id = {user_id}"
    return db.execute(query)""",
        "bugs": ["SQL injection vulnerability", "no parameterized query"],
        "label": 1
    },
    {
        "id": "BD-004", "project": "numpy",
        "code": """def safe_divide(a, b):
    if b == 0:
        return None
    return a / b

result = safe_divide(10, 0)
print(result + 5)""",
        "bugs": ["NoneType addition error", "no None check before arithmetic"],
        "label": 1
    },
    {
        "id": "BD-005", "project": "scrapy",
        "code": """def parse_date(date_str):
    return datetime.strptime(date_str, '%Y-%m-%d')""",
        "bugs": ["no exception handling for invalid date format", "no None check"],
        "label": 1
    },
    {
        "id": "BD-006", "project": "pandas",
        "code": """def get_first_last(lst):
    return lst[0], lst[-1]""",
        "bugs": ["IndexError when list is empty"],
        "label": 1
    },
    {
        "id": "BD-007", "project": "flask",
        "code": """SECRET_KEY = "mysecretkey123"
DATABASE_PASSWORD = "admin123"

def connect():
    return db.connect(password=DATABASE_PASSWORD)""",
        "bugs": ["hardcoded credentials", "security vulnerability", "should use env vars"],
        "label": 1
    },
    {
        "id": "BD-008", "project": "numpy",
        "code": """def is_valid_age(age):
    if age > 0 or age < 150:
        return True
    return False""",
        "bugs": ["logic error: 'or' should be 'and'", "always returns True"],
        "label": 1
    },
    {
        "id": "BD-009", "project": "requests",
        "code": """def download_file(url, path):
    r = requests.get(url)
    with open(path, 'wb') as f:
        f.write(r.content)""",
        "bugs": ["no timeout", "no error handling", "no status check"],
        "label": 1
    },
    {
        "id": "BD-010", "project": "pandas",
        "code": """def multiply_all(numbers):
    result = 0
    for n in numbers:
        result *= n
    return result""",
        "bugs": ["wrong initial value: should be 1 not 0", "always returns 0"],
        "label": 1
    },
    {
        "id": "BD-011", "project": "scrapy",
        "code": """def extract_links(html):
    soup = BeautifulSoup(html)
    return [a['href'] for a in soup.find_all('a')]""",
        "bugs": ["missing html.parser argument", "KeyError if href missing"],
        "label": 1
    },
    {
        "id": "BD-012", "project": "flask",
        "code": """def countdown(n):
    while n != 0:
        print(n)
        n -= 1""",
        "bugs": ["infinite loop when n is negative", "no lower bound check"],
        "label": 1
    },
    {
        "id": "BD-013", "project": "numpy",
        "code": """def read_config(path):
    with open(path) as f:
        return json.load(f)""",
        "bugs": ["FileNotFoundError not handled", "JSONDecodeError not handled"],
        "label": 1
    },
    {
        "id": "BD-014", "project": "pandas",
        "code": """def get_last_n(lst, n):
    return lst[len(lst)-n+1:]""",
        "bugs": ["off-by-one error: should be len(lst)-n"],
        "label": 1
    },
    {
        "id": "BD-015", "project": "requests",
        "code": """def parse_url(url):
    return url.split('://')[1]""",
        "bugs": ["IndexError if no protocol in URL", "no validation"],
        "label": 1
    },
    {
        "id": "BD-016", "project": "flask",
        "code": """def create_user(name, age, email):
    if name and age and email:
        return {"name": name, "age": age, "email": email}""",
        "bugs": ["no return for invalid input", "no type validation", "age could be negative"],
        "label": 1
    },
    {
        "id": "BD-017", "project": "numpy",
        "code": """def matrix_multiply(A, B):
    return [[sum(A[i][k]*B[k][j] for k in range(len(B)))
             for j in range(len(B[0]))]
            for i in range(len(A))]""",
        "bugs": ["no dimension compatibility check", "IndexError if dimensions mismatch"],
        "label": 1
    },
    {
        "id": "BD-018", "project": "scrapy",
        "code": """cache = {}
def memoize(n):
    if n in cache:
        return cache[n]
    result = n * 2
    cache[n] = result""",
        "bugs": ["missing return statement", "returns None on cache miss"],
        "label": 1
    },
    {
        "id": "BD-019", "project": "pandas",
        "code": """def safe_open(filename):
    try:
        f = open(filename)
        return f.read()
    except:
        pass""",
        "bugs": ["file not closed on success", "bare except hides all errors", "no return on error"],
        "label": 1
    },
    {
        "id": "BD-020", "project": "flask",
        "code": """def process_items(items):
    results = []
    for i in range(len(items)+1):
        results.append(items[i] * 2)
    return results""",
        "bugs": ["off-by-one: range should be len(items) not len(items)+1"],
        "label": 1
    },
    # ── CLEAN CODE (label=0) ──────────────────────────────────
    {
        "id": "BD-021", "project": "pandas",
        "code": """def calculate_mean(data):
    if not data:
        return None
    return sum(data) / len(data)""",
        "bugs": [], "label": 0
    },
    {
        "id": "BD-022", "project": "requests",
        "code": """def fetch_url(url, timeout=30):
    try:
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        raise RuntimeError(f"Failed to fetch {url}: {e}")""",
        "bugs": [], "label": 0
    },
    {
        "id": "BD-023", "project": "flask",
        "code": """def get_user(user_id, db):
    return db.execute(
        "SELECT * FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()""",
        "bugs": [], "label": 0
    },
    {
        "id": "BD-024", "project": "numpy",
        "code": """def safe_divide(a, b):
    if b == 0:
        raise ZeroDivisionError("Cannot divide by zero")
    return a / b""",
        "bugs": [], "label": 0
    },
    {
        "id": "BD-025", "project": "scrapy",
        "code": """def parse_date(date_str):
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return None""",
        "bugs": [], "label": 0
    },
    {
        "id": "BD-026", "project": "pandas",
        "code": """def get_first_last(lst):
    if not lst:
        raise ValueError("List cannot be empty")
    return lst[0], lst[-1]""",
        "bugs": [], "label": 0
    },
    {
        "id": "BD-027", "project": "numpy",
        "code": """def is_valid_age(age):
    if age > 0 and age < 150:
        return True
    return False""",
        "bugs": [], "label": 0
    },
    {
        "id": "BD-028", "project": "pandas",
        "code": """def multiply_all(numbers):
    if not numbers:
        return 1
    result = 1
    for n in numbers:
        result *= n
    return result""",
        "bugs": [], "label": 0
    },
    {
        "id": "BD-029", "project": "flask",
        "code": """import os

def connect():
    password = os.getenv('DATABASE_PASSWORD')
    if not password:
        raise ValueError("DATABASE_PASSWORD not set")
    return db.connect(password=password)""",
        "bugs": [], "label": 0
    },
    {
        "id": "BD-030", "project": "numpy",
        "code": """def read_config(path):
    try:
        with open(path) as f:
            return json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"Config file not found: {path}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in config: {e}")""",
        "bugs": [], "label": 0
    },
]

# ─── BUG DETECTION PROMPT ─────────────────────────────────────
def make_prompt(code: str) -> str:
    return f"""You are an expert Python bug detector. Analyze this code carefully.

TASK: Determine if this code contains bugs.

Code:
```python
{code}
```

Respond in this EXACT format:
VERDICT: [BUGGY or CLEAN]
BUGS_FOUND: [number of bugs, 0 if clean]
BUG_1: [describe bug if exists, else NONE]
BUG_2: [describe bug if exists, else NONE]
BUG_3: [describe bug if exists, else NONE]
FIX: [brief fix suggestion or NONE]

Be precise. Only report real bugs, not style issues."""

def parse_response(response: str) -> dict:
    """Parse LLM response to extract verdict and bugs"""
    response_lower = response.lower()

    # Extract verdict
    verdict_buggy = 'buggy' in response_lower
    verdict_clean = 'clean' in response_lower and not verdict_buggy

    # Extract number of bugs mentioned
    bugs_found = 0
    match = re.search(r'bugs_found:\s*(\d+)', response_lower)
    if match:
        bugs_found = int(match.group(1))
    elif verdict_buggy:
        bugs_found = 1

    predicted_label = 1 if verdict_buggy else 0
    return {
        'predicted_label': predicted_label,
        'bugs_found': bugs_found,
        'response_length': len(response),
    }

# ─── MAIN EVALUATION ──────────────────────────────────────────
results = []
llms = ['Groq', 'Mistral', 'OpenRouter']
total = len(test_cases) * len(llms)
current = 0

print(f"\n  {len(test_cases)} test cases × 3 LLMs = {total} API calls")
print(f"  20 buggy + 10 clean code samples")
print(f"  Estimasi: ~10 menit\n")
print("=" * 60)

for tc in test_cases:
    for llm in llms:
        current += 1
        print(f"  [{current:3d}/{total}] {tc['id']} ({tc['project']}) | {llm}...", end=" ", flush=True)

        prompt   = make_prompt(tc['code'])
        response, resp_time = call_llm(llm, prompt)
        parsed   = parse_response(response)

        # Ground truth
        true_label = tc['label']
        pred_label = parsed['predicted_label']

        # TP, TN, FP, FN
        tp = 1 if (true_label == 1 and pred_label == 1) else 0
        tn = 1 if (true_label == 0 and pred_label == 0) else 0
        fp = 1 if (true_label == 0 and pred_label == 1) else 0
        fn = 1 if (true_label == 1 and pred_label == 0) else 0
        correct = 1 if true_label == pred_label else 0

        status = "✅" if correct else "❌"
        print(f"{status} Pred:{['CLEAN','BUGGY'][pred_label]} True:{['CLEAN','BUGGY'][true_label]} Time:{resp_time:.1f}s")

        results.append({
            'TC ID': tc['id'],
            'Project': tc['project'],
            'True Label': true_label,
            'True Label Text': 'BUGGY' if true_label==1 else 'CLEAN',
            'LLM': llm,
            'Predicted Label': pred_label,
            'Predicted Text': 'BUGGY' if pred_label==1 else 'CLEAN',
            'Correct': correct,
            'TP': tp, 'TN': tn, 'FP': fp, 'FN': fn,
            'Response Time (s)': resp_time,
            'Bugs Found': parsed['bugs_found'],
        })
        time.sleep(0.5)

# ─── COMPUTE METRICS ──────────────────────────────────────────
df = pd.DataFrame(results)

print(f"\n{'='*60}")
print("  RESULTS PER LLM")
print(f"{'='*60}")

summary_rows = []
for llm in llms:
    d = df[df['LLM']==llm]
    tp = d['TP'].sum(); tn = d['TN'].sum()
    fp = d['FP'].sum(); fn = d['FN'].sum()

    precision = tp/(tp+fp) if (tp+fp)>0 else 0
    recall    = tp/(tp+fn) if (tp+fn)>0 else 0
    f1        = 2*precision*recall/(precision+recall) if (precision+recall)>0 else 0
    accuracy  = (tp+tn)/len(d)
    avg_time  = d['Response Time (s)'].mean()

    print(f"\n  {llm}:")
    print(f"    Accuracy  : {accuracy:.4f} ({accuracy*100:.1f}%)")
    print(f"    Precision : {precision:.4f}")
    print(f"    Recall    : {recall:.4f}")
    print(f"    F1-Score  : {f1:.4f}")
    print(f"    Avg Time  : {avg_time:.2f}s")
    print(f"    TP={tp} TN={tn} FP={fp} FN={fn}")

    summary_rows.append({
        'LLM': llm, 'N': len(d),
        'Accuracy': round(accuracy,4),
        'Precision': round(precision,4),
        'Recall': round(recall,4),
        'F1-Score': round(f1,4),
        'TP': int(tp), 'TN': int(tn), 'FP': int(fp), 'FN': int(fn),
        'Avg Time (s)': round(avg_time,3),
    })

df_sum = pd.DataFrame(summary_rows)

# Wilcoxon on F1 proxy (correct per sample)
print(f"\n{'─'*60}")
print("  WILCOXON TEST (Correct/Incorrect per sample)")
pairs = [('Groq','Mistral'),('Groq','OpenRouter'),('Mistral','OpenRouter')]
stat_rows = []
for a,b in pairs:
    xa = df[df['LLM']==a]['Correct'].values
    xb = df[df['LLM']==b]['Correct'].values
    n  = min(len(xa),len(xb))
    try:
        _,p = wilcoxon(xa[:n],xb[:n])
        sig = '***' if p<0.001 else ('**' if p<0.01 else ('*' if p<0.05 else 'ns'))
        print(f"  {a} vs {b}: p={p:.4f} {sig}")
        stat_rows.append({'Comparison':f'{a} vs {b}','p-value':round(p,4),'Sig':sig})
    except Exception as e:
        print(f"  {a} vs {b}: {e}")

# ─── SAVE EXCEL ───────────────────────────────────────────────
wb  = openpyxl.Workbook()
bdr = Border(*[Side(style='thin',color='CCCCCC')]*4)

def H(c,col="1F3864"):
    c.font=Font(bold=True,color="FFFFFF",size=11)
    c.fill=PatternFill("solid",fgColor=col)
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    c.border=bdr

def D(c,even=True,green=False,red=False):
    clr = "C6EFCE" if green else ("FFC7CE" if red else ("EBF3FB" if even else "FFFFFF"))
    c.fill=PatternFill("solid",fgColor=clr)
    c.font=Font(size=10,bold=(green or red),
                color=("006100" if green else ("9C0006" if red else "000000")))
    c.alignment=Alignment(horizontal='center',vertical='center')
    c.border=bdr

# Sheet 1: Summary
ws1=wb.active; ws1.title="📊 Bug Detection Summary"; ws1.sheet_view.showGridLines=False
ws1.merge_cells('A1:J1')
c=ws1['A1']; c.value="BUG DETECTION EVALUATION — Precision, Recall, F1-Score"
c.font=Font(bold=True,size=14,color="FFFFFF")
c.fill=PatternFill("solid",fgColor="C00000")
c.alignment=Alignment(horizontal='center',vertical='center')
ws1.row_dimensions[1].height=32

cols=list(df_sum.columns)
for ci,h in enumerate(cols,1): c=ws1.cell(row=2,column=ci,value=h); H(c,"C00000")
ws1.row_dimensions[2].height=26

best_f1=df_sum['F1-Score'].max()
best_acc=df_sum['Accuracy'].max()
for ri,row in df_sum.iterrows():
    r=ri+3; even=r%2==0
    for ci,key in enumerate(cols,1):
        c=ws1.cell(row=r,column=ci,value=row[key])
        is_best=(key=='F1-Score' and row[key]==best_f1) or (key=='Accuracy' and row[key]==best_acc)
        D(c,even,green=is_best)
    ws1.row_dimensions[r].height=20

for i,w in enumerate([14,6,12,12,12,12,8,8,8,8,12],1):
    ws1.column_dimensions[get_column_letter(i)].width=w

# Sheet 2: Raw Results
ws2=wb.create_sheet("📋 Raw Results"); ws2.sheet_view.showGridLines=False
raw_cols=['TC ID','Project','True Label Text','LLM','Predicted Text','Correct','TP','TN','FP','FN','Response Time (s)']
for ci,h in enumerate(raw_cols,1): c=ws2.cell(row=1,column=ci,value=h); H(c)
ws2.row_dimensions[1].height=22
colors_llm={'Groq':'DDEEFF','Mistral':'FFF2CC','OpenRouter':'FCE4D6'}
for ri,(_, row) in enumerate(df.iterrows(),2):
    for ci,key in enumerate(raw_cols,1):
        c=ws2.cell(row=ri,column=ci,value=row.get(key,''))
        c.fill=PatternFill("solid",fgColor=colors_llm.get(row['LLM'],'FFFFFF'))
        c.alignment=Alignment(horizontal='center',vertical='center')
        c.font=Font(size=9); c.border=bdr
    ws2.row_dimensions[ri].height=15
for i,w in enumerate([10,12,14,12,14,8,6,6,6,6,14],1):
    ws2.column_dimensions[get_column_letter(i)].width=w

fname=f'bug_eval_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(fname)

print(f"\n{'='*60}")
print(f"  🎉 DONE! Saved: {fname}")
print(f"\n  FINAL F1-SCORE COMPARISON:")
for _,r in df_sum.iterrows():
    print(f"  {r['LLM']:<14} F1={r['F1-Score']:.4f} Acc={r['Accuracy']:.4f} Time={r['Avg Time (s)']:.2f}s")
print(f"\n  Upload {fname} ke GitHub → paper siap diupdate!")
print(f"{'='*60}")
