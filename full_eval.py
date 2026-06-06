"""
=====================================================
FULL EVALUATION — LLM Code Intelligence Research
164 HumanEval Test Cases × 3 LLMs
Metrics: BLEU, ROUGE, Pass@1, Wilcoxon, Effect Size
=====================================================
Cara pakai:
1. pip install groq mistralai openai rouge-score nltk scipy openpyxl pandas python-dotenv
2. python full_eval.py
3. Hasil: full_eval_results.xlsx (siap untuk paper Q1)

Estimasi waktu: 30-60 menit
"""

import os
import time
import json
import subprocess
import tempfile
import pandas as pd
import numpy as np
from dotenv import load_dotenv
from datetime import datetime
from pathlib import Path

# Metrics
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
from rouge_score import rouge_scorer
from scipy.stats import wilcoxon
from scipy.stats import rankdata
import warnings
warnings.filterwarnings('ignore')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

print("=" * 65)
print("  FULL EVALUATION — LLM Code Intelligence")
print("  164 HumanEval × 3 LLMs (Groq + Mistral + OpenRouter)")
print("=" * 65)

# ─── API CLIENTS ──────────────────────────────────────────────
from groq import Groq
from mistralai import Mistral
from openai import OpenAI

groq_client    = Groq(api_key=os.getenv("GROQ_API_KEY"))
mistral_client = Mistral(api_key=os.getenv("MISTRAL_API_KEY"))
openrouter_client = OpenAI(
    api_key=os.getenv("OPENROUTER_API_KEY"),
    base_url="https://openrouter.ai/api/v1"
)

# ─── LLM CALL FUNCTIONS ───────────────────────────────────────
def call_groq(prompt: str) -> tuple[str, float]:
    try:
        t0 = time.time()
        res = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1024
        )
        return res.choices[0].message.content, round(time.time()-t0, 3)
    except Exception as e:
        return f"ERROR: {e}", 0.0

def call_mistral(prompt: str) -> tuple[str, float]:
    try:
        t0 = time.time()
        res = mistral_client.chat.complete(
            model="mistral-small-latest",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1024
        )
        return res.choices[0].message.content, round(time.time()-t0, 3)
    except Exception as e:
        return f"ERROR: {e}", 0.0

def call_openrouter(prompt: str) -> tuple[str, float]:
    try:
        t0 = time.time()
        res = openrouter_client.chat.completions.create(
            model="openai/gpt-oss-120b:free",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1024
        )
        return res.choices[0].message.content, round(time.time()-t0, 3)
    except Exception as e:
        return f"ERROR: {e}", 0.0

# ─── METRIC FUNCTIONS ─────────────────────────────────────────
rouge_sc = rouge_scorer.RougeScorer(['rouge1','rouge2','rougeL'], use_stemmer=True)
smooth   = SmoothingFunction().method1

def calc_bleu(reference: str, hypothesis: str) -> float:
    if not hypothesis or not reference:
        return 0.0
    try:
        ref_tokens = [reference.lower().split()]
        hyp_tokens = hypothesis.lower().split()
        if len(hyp_tokens) < 1:
            return 0.0
        return round(sentence_bleu(ref_tokens, hyp_tokens, smoothing_function=smooth), 4)
    except:
        return 0.0

def calc_rouge(reference: str, hypothesis: str) -> dict:
    if not hypothesis or not reference:
        return {'r1': 0.0, 'r2': 0.0, 'rL': 0.0}
    try:
        s = rouge_sc.score(reference, hypothesis)
        return {
            'r1': round(s['rouge1'].fmeasure, 4),
            'r2': round(s['rouge2'].fmeasure, 4),
            'rL': round(s['rougeL'].fmeasure, 4),
        }
    except:
        return {'r1': 0.0, 'r2': 0.0, 'rL': 0.0}

def extract_python_code(text: str) -> str:
    """Extract Python code from LLM response"""
    import re
    # Try ```python blocks first
    pattern = r'```python\n(.*?)```'
    matches = re.findall(pattern, text, re.DOTALL)
    if matches:
        return matches[0].strip()
    # Try ``` blocks
    pattern2 = r'```\n?(.*?)```'
    matches2 = re.findall(pattern2, text, re.DOTALL)
    if matches2:
        return matches2[0].strip()
    # Return as-is
    return text.strip()

def run_pass_at_1(generated_code: str, test_code: str, entry_point: str) -> int:
    """
    Run Pass@1: execute generated function + test
    Returns 1 if passes, 0 if fails
    """
    try:
        full_code = f"{generated_code}\n\n{test_code}"
        with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False) as f:
            f.write(full_code)
            fname = f.name

        result = subprocess.run(
            ['python', fname],
            capture_output=True, text=True, timeout=10
        )
        os.unlink(fname)
        return 1 if result.returncode == 0 else 0
    except:
        return 0

def rank_biserial_r(x, y):
    """Effect size: rank-biserial correlation"""
    try:
        n1, n2 = len(x), len(y)
        combined = np.concatenate([x, y])
        ranks = rankdata(combined)
        r1 = np.sum(ranks[:n1])
        U1 = r1 - n1*(n1+1)/2
        r = 1 - (2*U1)/(n1*n2)
        return round(r, 4)
    except:
        return 0.0

# ─── LOAD HUMANEVAL DATASET ───────────────────────────────────
print("\n📦 Loading HumanEval dataset...")
df_he = pd.read_csv('data/humaneval/humaneval.csv')
print(f"   ✅ {len(df_he)} problems loaded")
print(f"   Columns: {list(df_he.columns)}")

# Use first N for faster testing (set to len(df_he) for full run)
# For paper: use all 164
N_SAMPLES = len(df_he)  # Change to e.g. 20 for quick test
df_eval = df_he.head(N_SAMPLES).reset_index(drop=True)

print(f"\n   Running on {N_SAMPLES} test cases...")
print(f"   Estimated time: ~{N_SAMPLES * 3 * 3 / 60:.0f} minutes")

# ─── PROMPTS ──────────────────────────────────────────────────
def make_test_gen_prompt(problem: dict) -> str:
    return f"""You are a Python testing expert. Generate comprehensive pytest unit tests for this function.

Function specification:
{problem.get('prompt', '')}

Entry point: {problem.get('entry_point', 'solution')}

Requirements:
1. Write complete pytest test functions
2. Cover: happy path, edge cases, error cases
3. Use assert statements
4. Return ONLY the test code, no explanation

Generate the test code:"""

def make_code_gen_prompt(problem: dict) -> str:
    return f"""Complete this Python function. Return ONLY the implementation code, no explanation.

{problem.get('prompt', '')}

Implementation:"""

# ─── MAIN EVALUATION LOOP ─────────────────────────────────────
results = []
llms = ['Groq', 'Mistral', 'OpenRouter']
callers = {
    'Groq': call_groq,
    'Mistral': call_mistral,
    'OpenRouter': call_openrouter
}

total = N_SAMPLES * len(llms)
current = 0

print("\n" + "=" * 65)
print("  STARTING EVALUATION...")
print("  (Ctrl+C to stop early — partial results will be saved)")
print("=" * 65)

save_checkpoint = []

try:
    for idx, row in df_eval.iterrows():
        tc_id = row.get('task_id', f'HE-{idx:03d}')
        prompt_text = row.get('prompt', '')
        canonical = row.get('canonical_solution', '')
        test_code = row.get('test', '')
        entry_point = row.get('entry_point', 'solution')

        # Build reference (canonical solution as ground truth for BLEU/ROUGE)
        reference_solution = f"{prompt_text}{canonical}"

        for llm in llms:
            current += 1
            print(f"\n  [{current:3d}/{total}] {tc_id} | {llm}...", end=" ", flush=True)

            # ── Task: Test Generation ──────────────────────────
            tg_prompt = make_test_gen_prompt(row.to_dict())
            tg_response, tg_time = callers[llm](tg_prompt)

            # BLEU on test generation
            tg_bleu = calc_bleu(test_code, tg_response)
            tg_rouge = calc_rouge(test_code, tg_response)

            # ── Task: Code Generation (for Pass@1) ────────────
            cg_prompt = make_code_gen_prompt(row.to_dict())
            cg_response, cg_time = callers[llm](cg_prompt)

            # Extract code and run Pass@1
            extracted_code = extract_python_code(cg_response)
            full_fn = f"{prompt_text}{extracted_code}"
            pass1 = run_pass_at_1(full_fn, test_code, entry_point)

            # BLEU/ROUGE on code generation
            cg_bleu = calc_bleu(canonical, extracted_code)
            cg_rouge = calc_rouge(canonical, extracted_code)

            print(f"✅ TG-BLEU:{tg_bleu:.3f} | Pass@1:{pass1} | Time:{tg_time:.1f}s")

            result = {
                'Task ID': tc_id,
                'Entry Point': entry_point,
                'LLM': llm,
                # Test Generation metrics
                'TG Response Time (s)': tg_time,
                'TG BLEU': tg_bleu,
                'TG ROUGE-1': tg_rouge['r1'],
                'TG ROUGE-2': tg_rouge['r2'],
                'TG ROUGE-L': tg_rouge['rL'],
                # Code Generation metrics
                'CG Response Time (s)': cg_time,
                'CG BLEU': cg_bleu,
                'CG ROUGE-L': cg_rouge['rL'],
                'Pass@1': pass1,
                # Raw responses (for transparency)
                'TG Response (excerpt)': tg_response[:200],
                'CG Response (excerpt)': cg_response[:200],
            }
            results.append(result)
            save_checkpoint.append(result)

            # Rate limit protection
            time.sleep(0.5)

        # Save checkpoint every 10 problems
        if (idx + 1) % 10 == 0:
            df_cp = pd.DataFrame(save_checkpoint)
            df_cp.to_csv('eval_checkpoint.csv', index=False)
            print(f"\n  💾 Checkpoint saved ({len(save_checkpoint)} records)")

except KeyboardInterrupt:
    print("\n\n⚠️  Evaluation stopped by user. Saving partial results...")

# ─── COMPUTE SUMMARY STATISTICS ───────────────────────────────
print("\n" + "=" * 65)
print("  COMPUTING STATISTICS...")
print("=" * 65)

df_results = pd.DataFrame(results)

if len(df_results) == 0:
    print("❌ No results to analyze. Exiting.")
    exit()

summary_rows = []
metrics_list = ['TG BLEU', 'TG ROUGE-L', 'CG BLEU', 'CG ROUGE-L', 'Pass@1',
                'TG Response Time (s)', 'CG Response Time (s)']

for metric in metrics_list:
    print(f"\n📊 {metric}:")
    llm_data = {}
    for llm in llms:
        vals = df_results[df_results['LLM'] == llm][metric].dropna().values
        llm_data[llm] = vals
        print(f"   {llm:<14} mean={np.mean(vals):.4f}  std={np.std(vals):.4f}  n={len(vals)}")

    # Wilcoxon tests between pairs
    pairs = [('Groq','Mistral'), ('Groq','OpenRouter'), ('Mistral','OpenRouter')]
    for a, b in pairs:
        try:
            xa, xb = llm_data[a], llm_data[b]
            min_len = min(len(xa), len(xb))
            if min_len >= 10:
                stat, pval = wilcoxon(xa[:min_len], xb[:min_len])
                r = rank_biserial_r(xa[:min_len], xb[:min_len])
                sig = "***" if pval < 0.001 else ("**" if pval < 0.01 else ("*" if pval < 0.05 else "ns"))
                print(f"   {a} vs {b}: p={pval:.4f} {sig} | r={r:.3f}")

                summary_rows.append({
                    'Metric': metric,
                    'LLM A': a, 'LLM B': b,
                    f'{a} Mean': round(np.mean(xa), 4),
                    f'{b} Mean': round(np.mean(xb), 4),
                    'Wilcoxon p': round(pval, 4),
                    'Significance': sig,
                    'Effect Size r': r,
                })
        except Exception as e:
            print(f"   {a} vs {b}: insufficient data")

# ─── AGGREGATE SUMMARY PER LLM ────────────────────────────────
print("\n" + "=" * 65)
print("  FINAL SUMMARY PER LLM")
print("=" * 65)

agg_rows = []
for llm in llms:
    d = df_results[df_results['LLM'] == llm]
    row = {
        'LLM': llm,
        'N': len(d),
        'Avg TG BLEU': round(d['TG BLEU'].mean(), 4),
        'Avg TG ROUGE-L': round(d['TG ROUGE-L'].mean(), 4),
        'Avg CG BLEU': round(d['CG BLEU'].mean(), 4),
        'Avg CG ROUGE-L': round(d['CG ROUGE-L'].mean(), 4),
        'Pass@1 Rate': round(d['Pass@1'].mean(), 4),
        'Avg TG Time (s)': round(d['TG Response Time (s)'].mean(), 3),
        'Avg CG Time (s)': round(d['CG Response Time (s)'].mean(), 3),
    }
    agg_rows.append(row)
    print(f"\n  {llm}")
    for k, v in row.items():
        if k != 'LLM':
            print(f"    {k:<22}: {v}")

df_agg = pd.DataFrame(agg_rows)

# ─── SAVE TO EXCEL ────────────────────────────────────────────
print("\n💾 Saving to Excel...")

wb = openpyxl.Workbook()
thin = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def hdr(cell, color="1F3864"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin

def dat(cell, even=True, green=False, red=False):
    if green:
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(bold=True, color="006100", size=10)
    elif red:
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        cell.font = Font(bold=True, color="9C0006", size=10)
    else:
        cell.fill = PatternFill("solid", fgColor="EBF3FB" if even else "FFFFFF")
        cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin

# ── Sheet 1: Aggregate Summary ──
ws1 = wb.active
ws1.title = "📊 Summary"
ws1.sheet_view.showGridLines = False

ws1.merge_cells('A1:J1')
c = ws1['A1']
c.value = f"LLM CODE INTELLIGENCE — FULL EVALUATION ({N_SAMPLES} HumanEval Problems)"
c.font = Font(bold=True, size=14, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F3864")
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 32

# Find best per metric
best_vals = {}
for col in ['Avg TG BLEU','Avg TG ROUGE-L','Avg CG BLEU','Avg CG ROUGE-L','Pass@1 Rate']:
    best_vals[col] = df_agg[col].max()
for col in ['Avg TG Time (s)','Avg CG Time (s)']:
    best_vals[col] = df_agg[col].min()

cols_agg = list(df_agg.columns)
for ci, h in enumerate(cols_agg, 1):
    c = ws1.cell(row=2, column=ci, value=h)
    hdr(c, "2E75B6")
ws1.row_dimensions[2].height = 28

for ri, row in df_agg.iterrows():
    r = ri + 3
    even = r % 2 == 0
    for ci, key in enumerate(cols_agg, 1):
        val = row[key]
        c = ws1.cell(row=r, column=ci, value=val)
        is_best = False
        if key in best_vals:
            is_best = (val == best_vals[key])
        dat(c, even, green=is_best)
    ws1.row_dimensions[r].height = 20

for i, w in enumerate([16,8,14,14,14,14,12,14,14], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Add interpretation note
note_row = len(df_agg) + 5
ws1.cell(row=note_row, column=1, value="📌 Interpretation: Green = Best per metric | Pass@1 = % of generated code that passes HumanEval test suite")
ws1.cell(row=note_row, column=1).font = Font(italic=True, color="2E75B6", size=10)
ws1.merge_cells(f'A{note_row}:J{note_row}')

# ── Sheet 2: Wilcoxon Statistical Tests ──
ws2 = wb.create_sheet("📈 Statistical Tests")
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:H1')
c = ws2['A1']
c.value = "WILCOXON SIGNED-RANK TEST — Pairwise LLM Comparison"
c.font = Font(bold=True, size=13, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F3864")
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

if summary_rows:
    df_stat = pd.DataFrame(summary_rows)
    stat_cols = list(df_stat.columns)
    for ci, h in enumerate(stat_cols, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        hdr(c, "7030A0")
    ws2.row_dimensions[2].height = 25

    for ri, row in df_stat.iterrows():
        r = ri + 3
        even = r % 2 == 0
        for ci, key in enumerate(stat_cols, 1):
            val = row[key]
            c = ws2.cell(row=r, column=ci, value=val)
            is_sig = (key == 'Significance' and val in ['*','**','***'])
            dat(c, even, green=is_sig)
        ws2.row_dimensions[r].height = 18

    for i, w in enumerate([22,10,10,12,12,12,12,12],1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Add significance legend
    leg_row = len(df_stat) + 4
    ws2.cell(row=leg_row, column=1, value="Significance: *** p<0.001 | ** p<0.01 | * p<0.05 | ns = not significant")
    ws2.cell(row=leg_row, column=1).font = Font(italic=True, color="7030A0", size=10)
    ws2.merge_cells(f'A{leg_row}:H{leg_row}')
    ws2.cell(row=leg_row+1, column=1, value="Effect Size r: |r|>0.5 = large | |r|>0.3 = medium | |r|>0.1 = small")
    ws2.cell(row=leg_row+1, column=1).font = Font(italic=True, color="7030A0", size=10)
    ws2.merge_cells(f'A{leg_row+1}:H{leg_row+1}')

# ── Sheet 3: Raw Results ──
ws3 = wb.create_sheet("📋 Raw Results")
ws3.sheet_view.showGridLines = False

raw_cols = ['Task ID','LLM','TG BLEU','TG ROUGE-L','CG BLEU','CG ROUGE-L',
            'Pass@1','TG Response Time (s)','CG Response Time (s)']
for ci, h in enumerate(raw_cols, 1):
    c = ws3.cell(row=1, column=ci, value=h)
    hdr(c)
ws3.row_dimensions[1].height = 22

llm_colors = {'Groq':'EBF3FB','Mistral':'FFF2CC','OpenRouter':'FCE4D6'}
for ri, row in df_results.iterrows():
    r = ri + 2
    llm = row['LLM']
    for ci, key in enumerate(raw_cols, 1):
        c = ws3.cell(row=r, column=ci, value=row.get(key,''))
        c.fill = PatternFill("solid", fgColor=llm_colors.get(llm,'FFFFFF'))
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(size=9)
        c.border = thin
    ws3.row_dimensions[r].height = 15

for i, w in enumerate([16,12,10,10,10,10,8,16,16],1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Save
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
fname = f'full_eval_results_{timestamp}.xlsx'
wb.save(fname)

print(f"\n{'='*65}")
print(f"  🎉 EVALUATION COMPLETE!")
print(f"  📊 Results: {fname}")
print(f"  📋 Total records: {len(df_results)}")
print(f"  🔬 Problems evaluated: {N_SAMPLES}")
print(f"  🤖 LLMs compared: {', '.join(llms)}")
print(f"\n  Key findings:")
for _, row in df_agg.iterrows():
    print(f"  {row['LLM']:<14} Pass@1={row['Pass@1 Rate']:.3f} | TG-BLEU={row['Avg TG BLEU']:.3f} | Avg Time={row['Avg TG Time (s)']:.2f}s")
print(f"{'='*65}")
print(f"\n  Upload {fname} ke GitHub dan masukkan ke paper!")
